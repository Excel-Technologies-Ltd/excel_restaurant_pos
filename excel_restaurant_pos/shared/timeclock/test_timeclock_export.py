# Copyright (c) 2026, Excel and Contributors
# See license.txt

import datetime
import os
import tempfile
from unittest.mock import patch

import frappe
from frappe.tests.utils import FrappeTestCase

from excel_restaurant_pos.shared.timeclock.export import (
	DEFAULT_COLUMNS,
	DOWNLOAD_TICKET_TTL,
	create_download_ticket,
	redeem_download_ticket,
	_cell_value,
	build_export_response,
	column_labels,
	iter_records,
	parse_filters,
	permitted_fieldnames,
	resolve_columns,
	write_workbook,
)

MODULE = "excel_restaurant_pos.shared.timeclock.export"


def _record(name, employee="6", business_date="2026-09-01"):
	return frappe._dict(
		{
			"name": name,
			"employee": employee,
			"employee_name": "Staff",
			"business_date": business_date,
			"first_check_in": f"{business_date} 09:00:00",
			"last_check_out": f"{business_date} 17:30:00",
			"total_paid_hours": 8.5,
			"timeclock_cost": 12.5,
			"total_payment": 106.25,
			"manual_entry": 0,
			"is_modified": 0,
			"modified_by_manager": None,
		}
	)


class TestExportArguments(FrappeTestCase):
	def test_timeclock_fields_are_permitted(self):
		permitted = permitted_fieldnames()
		for fieldname in DEFAULT_COLUMNS:
			self.assertIn(fieldname, permitted)

	def test_layout_fields_are_not_permitted(self):
		permitted = permitted_fieldnames()
		self.assertNotIn("timeclock_section", permitted)
		self.assertNotIn("column_break_main", permitted)

	def test_no_filters_exports_everything(self):
		self.assertEqual(parse_filters(None), [])
		self.assertEqual(parse_filters(""), [])
		self.assertEqual(parse_filters("[]"), [])

	def test_json_filters_are_parsed(self):
		filters = parse_filters('[["business_date", "between", ["2026-09-01", "2026-09-30"]]]')
		self.assertEqual(filters, [["business_date", "between", ["2026-09-01", "2026-09-30"]]])

	def test_dict_filters_are_accepted(self):
		self.assertEqual(parse_filters({"employee": "6"}), [["employee", "=", "6"]])

	def test_unknown_filter_field_is_rejected(self):
		with self.assertRaises(frappe.ValidationError):
			parse_filters('[["salary", "=", 1]]')

	def test_unsupported_operator_is_rejected(self):
		with self.assertRaises(frappe.ValidationError):
			parse_filters('[["employee", "; drop table x", 1]]')

	def test_malformed_filter_is_rejected(self):
		with self.assertRaises(frappe.ValidationError):
			parse_filters('[["employee", "="]]')

	def test_columns_default_to_the_full_record(self):
		self.assertEqual(resolve_columns(None), list(DEFAULT_COLUMNS))

	def test_unknown_column_is_rejected(self):
		with self.assertRaises(frappe.ValidationError):
			resolve_columns('["pin"]')

	def test_labels_come_from_the_doctype(self):
		self.assertEqual(
			column_labels(["business_date", "total_paid_hours"]),
			["Business Date", "Total Paid Hours"],
		)


class TestCellValues(FrappeTestCase):
	def test_check_fields_are_integers(self):
		self.assertEqual(_cell_value("1", "Check"), 1)
		self.assertEqual(_cell_value(None, "Check"), None)

	def test_currency_and_float_are_numbers(self):
		self.assertEqual(_cell_value("106.25", "Currency"), 106.25)
		self.assertEqual(_cell_value("8.5", "Float"), 8.5)

	def test_dates_are_native_objects(self):
		self.assertEqual(_cell_value("2026-09-01", "Date"), datetime.date(2026, 9, 1))
		self.assertEqual(
			_cell_value("2026-09-01 09:00:00", "Datetime"),
			datetime.datetime(2026, 9, 1, 9, 0, 0),
		)

	def test_blank_stays_blank(self):
		self.assertIsNone(_cell_value("", "Data"))
		self.assertIsNone(_cell_value(None, "Datetime"))


class TestRecordStreaming(FrappeTestCase):
	def test_rows_are_keyset_paged(self):
		pages = [
			[_record("ETT-2026-09-01-6"), _record("ETT-2026-09-01-7")],
			[_record("ETT-2026-09-02-6")],
		]

		with patch(f"{MODULE}.frappe.get_list", side_effect=pages) as get_list:
			names = [row["name"] for row in iter_records([], list(DEFAULT_COLUMNS), batch_size=2)]

		self.assertEqual(names, ["ETT-2026-09-01-6", "ETT-2026-09-01-7", "ETT-2026-09-02-6"])
		# Second page continues after the last name, never with an offset.
		second_call_filters = get_list.call_args_list[1].kwargs["filters"]
		self.assertIn(["name", ">", "ETT-2026-09-01-7"], second_call_filters)

	def test_short_page_ends_the_scan(self):
		with patch(f"{MODULE}.frappe.get_list", return_value=[_record("ETT-1")]) as get_list:
			list(iter_records([], list(DEFAULT_COLUMNS), batch_size=10))

		self.assertEqual(get_list.call_count, 1)

	def test_rows_are_read_with_the_callers_permissions(self):
		with patch(f"{MODULE}.frappe.get_list", return_value=[]) as get_list:
			list(iter_records([], list(DEFAULT_COLUMNS)))

		self.assertIs(get_list.call_args.kwargs["ignore_permissions"], False)

	def test_row_cap_is_enforced(self):
		with patch(f"{MODULE}.MAX_ROWS", 1):
			with patch(f"{MODULE}.frappe.get_list", return_value=[_record("A"), _record("B")]):
				with self.assertRaises(frappe.ValidationError):
					list(iter_records([], list(DEFAULT_COLUMNS), batch_size=2))


class TestWorkbook(FrappeTestCase):
	def test_workbook_has_a_header_and_one_row_per_record(self):
		import openpyxl

		handle, path = tempfile.mkstemp(suffix=".xlsx")
		os.close(handle)

		with patch(f"{MODULE}.frappe.get_list", return_value=[_record("ETT-2026-09-01-6")]):
			rows = write_workbook([], list(DEFAULT_COLUMNS), path)

		self.assertEqual(rows, 1)
		try:
			sheet = openpyxl.load_workbook(path).active
			self.assertEqual(sheet.max_row, 2)
			self.assertEqual(sheet.cell(row=1, column=1).value, "Name")
			self.assertEqual(sheet.cell(row=2, column=1).value, "ETT-2026-09-01-6")
			self.assertEqual(sheet.cell(row=2, column=4).number_format, "yyyy-mm-dd")
		finally:
			os.remove(path)


class TestExportResponse(FrappeTestCase):
	def test_export_permission_is_required(self):
		with patch(f"{MODULE}.frappe.has_permission", side_effect=frappe.PermissionError):
			with self.assertRaises(frappe.PermissionError):
				build_export_response()

	def test_response_is_a_streamed_attachment(self):
		with patch(f"{MODULE}.frappe.has_permission", return_value=True):
			with patch(f"{MODULE}.frappe.get_list", return_value=[_record("ETT-2026-09-01-6")]):
				with patch(f"{MODULE}._log_export"):
					response = build_export_response()

		self.assertEqual(
			response.mimetype,
			"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
		)
		self.assertIn("attachment", response.headers["Content-Disposition"])
		self.assertIn(".xlsx", response.headers["Content-Disposition"])
		self.assertEqual(response.headers["X-Row-Count"], "1")
		self.assertIn("no-store", response.headers["Cache-Control"])

		body = b"".join(response.response)
		self.assertEqual(len(body), int(response.headers["Content-Length"]))
		# The temporary workbook is removed once the stream is drained.
		self.assertTrue(body.startswith(b"PK"))

	def test_rate_limit_blocks_a_burst(self):
		frappe.cache().delete_value(f"arcpos:timeclock_export:{frappe.session.user}")
		outcomes = []

		with patch(f"{MODULE}.frappe.has_permission", return_value=True):
			with patch(f"{MODULE}.frappe.get_list", return_value=[]):
				with patch(f"{MODULE}._log_export"):
					for _attempt in range(8):
						try:
							response = build_export_response()
							b"".join(response.response)
							outcomes.append("ok")
						except frappe.ValidationError:
							outcomes.append("throttled")

		self.assertIn("throttled", outcomes)
		self.assertEqual(outcomes.count("ok"), 6)


class TestDownloadTicket(FrappeTestCase):
	"""A cross origin SPA cannot put a bearer token on a browser navigation."""

	def test_ticket_freezes_the_filters(self):
		with patch(f"{MODULE}.frappe.has_permission", return_value=True):
			ticket = create_download_ticket(raw_filters='[["employee", "=", "6"]]')

		payload = frappe.cache().get_value(f"arcpos:timeclock_export_ticket:{ticket['ticket']}")
		self.assertEqual(payload["filters"], [["employee", "=", "6"]])
		self.assertEqual(payload["user"], frappe.session.user)
		self.assertEqual(ticket["expires_in"], DOWNLOAD_TICKET_TTL)

	def test_ticket_requires_export_permission(self):
		with patch(f"{MODULE}.frappe.has_permission", side_effect=frappe.PermissionError):
			with self.assertRaises(frappe.PermissionError):
				create_download_ticket()

	def test_ticket_is_single_use(self):
		with patch(f"{MODULE}.frappe.has_permission", return_value=True):
			ticket = create_download_ticket()

			with patch(f"{MODULE}.frappe.get_list", return_value=[]):
				with patch(f"{MODULE}._log_export"):
					with patch(f"{MODULE}.frappe.set_user"):
						response = redeem_download_ticket(ticket["ticket"])
						b"".join(response.response)

			with self.assertRaises(frappe.AuthenticationError):
				redeem_download_ticket(ticket["ticket"])

	def test_forged_ticket_is_rejected(self):
		with self.assertRaises(frappe.AuthenticationError):
			redeem_download_ticket("not-a-real-ticket")

	def test_missing_ticket_is_rejected(self):
		with self.assertRaises(frappe.AuthenticationError):
			redeem_download_ticket(None)

	def test_redemption_runs_as_the_minting_user(self):
		with patch(f"{MODULE}.frappe.has_permission", return_value=True):
			ticket = create_download_ticket()

			with patch(f"{MODULE}.frappe.get_list", return_value=[]):
				with patch(f"{MODULE}._log_export"):
					with patch(f"{MODULE}.frappe.set_user") as set_user:
						response = redeem_download_ticket(ticket["ticket"])
						b"".join(response.response)

		set_user.assert_called_once_with(frappe.session.user)

	def test_cross_origin_headers_are_exposed(self):
		with patch(f"{MODULE}.frappe.has_permission", return_value=True):
			with patch(f"{MODULE}.frappe.get_list", return_value=[]):
				with patch(f"{MODULE}._log_export"):
					response = build_export_response()
					b"".join(response.response)

		exposed = response.headers["Access-Control-Expose-Headers"]
		for header in ("Content-Disposition", "Content-Length", "X-Row-Count"):
			self.assertIn(header, exposed)
		self.assertEqual(response.headers["Referrer-Policy"], "no-referrer")
